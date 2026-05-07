import streamlit as st
import pandas as pd
import google.generativeai as genai
from PIL import Image
import io
import base64
import json
import math
import re
import hashlib
from openpyxl import load_workbook
from streamlit_cookies_controller import CookieController

# 1. 페이지 기본 설정
st.set_page_config(page_title="노랑조명 명세서 변환기", layout="centered")


# ============================================================
# 비밀번호 잠금 (30일 자동 로그인)
# ============================================================
def check_password():
    """비밀번호 입력 화면. 맞으면 쿠키에 30일 저장하여 자동 로그인."""
    
    cookie_controller = CookieController(key="norang_auth_cookies")
    auth_password = st.secrets["APP_PASSWORD"]
    expected_hash = hashlib.sha256(auth_password.encode()).hexdigest()
    
    # 1. 이미 세션에서 인증됨
    if st.session_state.get("password_correct", False):
        return True
    
    # 2. 쿠키 자동 로그인 체크 (첫 로드 시 None 안전 처리)
    try:
        cookie_value = cookie_controller.get('norang_auth')
    except (TypeError, AttributeError, KeyError):
        cookie_value = None
    
    if cookie_value == expected_hash:
        st.session_state["password_correct"] = True
        return True
    
    # 3. 첫 렌더링 시 쿠키가 None일 수 있어 한 번 더 시도
    if cookie_value is None and not st.session_state.get("_cookie_check_done", False):
        st.session_state["_cookie_check_done"] = True
        st.rerun()
    
    # 4. 비밀번호 입력 화면
    def password_entered():
        if st.session_state["password_input"] == auth_password:
            st.session_state["password_correct"] = True
            try:
                cookie_controller.set('norang_auth', expected_hash, max_age=30*24*60*60)
            except Exception:
                pass  # 쿠키 저장 실패해도 로그인은 진행
            del st.session_state["password_input"]
        else:
            st.session_state["password_correct"] = False
    
    st.markdown("""
        <div style='text-align:center; margin-top:80px; margin-bottom:30px;'>
            <h2 style='color:#333; font-weight:700;'>🔒 노랑조명 명세서 변환기</h2>
            <p style='color:#888; font-size:14px; margin-top:10px;'>매장 직원 전용 - 비밀번호를 입력해 주세요</p>
        </div>
    """, unsafe_allow_html=True)
    
    st.text_input(
        "비밀번호",
        type="password",
        on_change=password_entered,
        key="password_input",
        label_visibility="collapsed",
        placeholder="비밀번호를 입력하세요"
    )
    
    if "password_correct" in st.session_state:
        if not st.session_state["password_correct"]:
            st.error("❌ 비밀번호가 틀렸습니다. 다시 시도해 주세요.")
    
    st.markdown("<div style='border-top:1px solid #eee;margin-top:50px;padding-top:20px;text-align:center;font-size:11px;color:#bbb;'>Developed & Managed by <b>Eugene UG</b></div>", unsafe_allow_html=True)
    
    return False


# 비밀번호 통과 못하면 앱 본체 실행 안 함
if not check_password():
    st.stop()


# --- CSS 스타일 ---
st.markdown("""
    <style>
    .block-container { 
        padding-top: 3rem !important; 
        max-width: 730px !important;
    }
    .stApp { background-color: white; }
    .main-title {
        text-align: center;
        font-size: 26px;
        font-weight: 800;
        margin: 15px 0 25px 0;
    }
    [data-testid="stFileUploaderDropzone"] {
        min-height: 450px; 
        border: 2px solid #FFD400 !important; 
        background-color: #FFFDF0 !important; 
        border-radius: 15px !important;
        display: flex;
        justify-content: center;
        align-items: center;
        transition: all 0.3s ease !important;
    }
    [data-testid="stFileUploaderDropzone"]::before,
    [data-testid="stFileUploaderDropzone"]::after {
        display: none !important;
        content: none !important;
    }
    [data-testid="stFileUploaderDropzone"]:hover {
        background-color: #FFD400 !important;
        border: 0 !important;
        cursor: pointer !important;
    }
    [data-testid="stFileUploaderDropzone"][aria-over="true"] {
        border: 0 !important;
        background-color: #FFF9D0 !important;
        box-shadow: none !important;
    }
    [data-testid="stFileUploaderDropzone"] * {
        color: transparent !important;
        border-color: transparent !important;
    }
    .custom-label-wrapper * {
        color: inherit !important;
        pointer-events: none;
    }
    [data-testid="stFileUploaderDropzoneInstructions"],
    [data-testid="stFileUploader"] label, 
    [data-testid="stFileUploader"] small,
    [data-testid="stFileUploader"] button {
        display: none !important;
    }
    [data-testid="stFileChip"],
    [data-testid="stFileChips"],
    [data-testid="stFileChipDeleteBtn"],
    [data-testid="stBaseButton-borderlessIcon"] {
        display: none !important;
    }
    /* 업로드된 이미지의 fullscreen 아이콘을 이미지 안쪽 우측 상단으로 */
    /* stFullScreenFrame은 건드리지 말 것 (fullscreen 모드 작동 방해) */
    .stElementContainer:has(img),
    [data-testid="stElementContainer"]:has(img) {
        position: relative !important;
    }
    .stElementContainer:has(img) [data-testid="stElementToolbar"],
    [data-testid="stElementContainer"]:has(img) [data-testid="stElementToolbar"] {
        position: absolute !important;
        top: 8px !important;
        right: 13px !important;
        z-index: 99999 !important;
        background-color: transparent !important;
        pointer-events: auto !important;
        opacity: 1 !important;
        visibility: visible !important;
    }
    [data-testid="stBaseButton-elementToolbar"] {
        background-color: rgba(0, 0, 0, 0.6) !important;
        color: white !important;
        border: none !important;
        border-radius: 4px !important;
        pointer-events: auto !important;
        cursor: pointer !important;
    }
    [data-testid="stBaseButton-elementToolbar"]:hover {
        background-color: rgba(0, 0, 0, 0.85) !important;
    }
    /* 데이터 표만 좌우로 넓게 - transform으로 강제 가운데 정렬 */
    .stElementContainer:has([data-testid="stDataEditor"]),
    .stElementContainer:has([data-testid="stDataFrame"]),
    div[data-testid="stElementContainer"]:has([data-testid="stDataEditor"]) {
        width: 1100px !important;
        max-width: 1100px !important;
        position: relative !important;
        left: 50% !important;
        transform: translateX(-50%) !important;
    }
    /* 부여 예정 코드 박스 wrapper를 표와 같은 너비(1100px)로 가운데 정렬 */
    .stElementContainer:has(.code-preview-wrapper) {
        width: 1100px !important;
        max-width: 1100px !important;
        position: relative !important;
        left: 50% !important;
        transform: translateX(-50%) !important;
    }
    /* 데이터 분석 시작 버튼 색상 (secondary 타입) */
    [data-testid="stButton"] button[kind="secondary"] {
        background-color: #FFD400 !important;
        border-color: #FFD400 !important;
        color: #333 !important;
    }
    [data-testid="stButton"] button[kind="secondary"]:hover {
        background-color: #E6BE00 !important;
        border-color: #E6BE00 !important;
        color: #333 !important;
    }
    /* 검토 완료 버튼 - 엑셀 초록색 (다운로드 버튼은 영향 없음) */
    [data-testid="stButton"] button[kind="primary"] {
        background-color: #217346 !important;
        border-color: #217346 !important;
        color: white !important;
    }
    [data-testid="stButton"] button[kind="primary"]:hover {
        background-color: #1a5a37 !important;
        border-color: #1a5a37 !important;
        color: white !important;
    }
    [data-testid="stButton"] button[kind="primary"]:active,
    [data-testid="stButton"] button[kind="primary"]:focus {
        background-color: #154a2c !important;
        border-color: #154a2c !important;
        color: white !important;
        box-shadow: 0 0 0 0.2rem rgba(33, 115, 70, 0.25) !important;
    }
    </style>
""", unsafe_allow_html=True)


# ============================================================
# 헬퍼 함수
# ============================================================

def get_image_base64(path):
    try:
        with open(path, "rb") as f:
            return base64.b64encode(f.read()).decode()
    except: 
        return None

def load_brand_data():
    try:
        with open('brand_data.json', 'r', encoding='utf-8') as f:
            return json.load(f)
    except FileNotFoundError:
        st.error("⚠️ brand_data.json 파일이 없습니다. app.py와 같은 폴더에 두세요.")
        st.stop()

def is_shipping_item(text):
    """택배비/운임/배송비 등 부대비용 항목인지 판별"""
    if not text:
        return False
    s = str(text).strip()
    
    # 부대비용 키워드 (품목이 아닌 비용 항목)
    shipping_keywords = [
        '선불택배', '착불택배', '선불', '착불',
        '선불택', '착불택', '선부택', '착부택',  # 손글씨 줄임말
        '택배비', '택배', '운임', '운송비', '운송',
        '배송비', '배송료', '배송',
        '화물비', '화물',
        '차비', '차량비', '운반비', '운반',
        '퀵서비스', '퀵',
        '용달', '용달비',
        '선물박스', '포장박스', '박스비', '포장비', '포장',  # OCR이 "선불택"을 "선물박스" 등으로 잘못 읽는 경우
    ]
    
    for kw in shipping_keywords:
        if kw in s:
            return True
    return False


def is_empty_item(item_name, abbr):
    """품목명이 약자만 남거나 비어있는지 판별 (의미 없는 행)"""
    if not item_name:
        return True
    s = str(item_name).strip()
    # 약자 + ")" 패턴만 남은 경우 (예: "CHL)", "YD) ", "CHL) /")
    if abbr:
        no_abbr = s.replace(f"{abbr})", "").strip()
        if not no_abbr or no_abbr in ['/', '//', '-', '_']:
            return True
    # ") "로 시작하고 의미있는 텍스트 없는 경우
    no_prefix = re.sub(r'^[\w]*\)\s*', '', s)
    if not no_prefix or len(no_prefix) < 2:
        return True
    return False


def calculate_sales_price(purchase_price):
    """매출단가: (매입가 × 1.5) × 1.1, 백원 단위 올림"""
    if purchase_price is None or pd.isna(purchase_price):
        return 0
    raw = float(purchase_price) * 1.5 * 1.1
    return int(math.ceil(raw / 100) * 100)


import requests as _requests
import time as _time


def _github_get_counter():
    """GitHub API로 counter.json 읽기. (current_value, sha) 반환"""
    token = st.secrets["GITHUB_TOKEN"]
    repo = st.secrets["GITHUB_REPO"]
    url = f"https://api.github.com/repos/{repo}/contents/counter.json"
    headers = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/vnd.github+json",
    }
    response = _requests.get(url, headers=headers, timeout=10)
    response.raise_for_status()
    data = response.json()
    
    # base64로 인코딩되어 있어 디코딩 필요
    import base64 as _b64
    content_decoded = _b64.b64decode(data['content']).decode('utf-8')
    counter_data = json.loads(content_decoded)
    current = counter_data.get('current', 1)
    sha = data['sha']
    return current, sha


def _github_update_counter(new_value, sha):
    """GitHub API로 counter.json 업데이트. 성공시 True 반환"""
    token = st.secrets["GITHUB_TOKEN"]
    repo = st.secrets["GITHUB_REPO"]
    url = f"https://api.github.com/repos/{repo}/contents/counter.json"
    headers = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/vnd.github+json",
    }
    
    import base64 as _b64
    new_content = json.dumps({"current": new_value})
    content_encoded = _b64.b64encode(new_content.encode('utf-8')).decode('utf-8')
    
    payload = {
        "message": f"Update counter to {new_value}",
        "content": content_encoded,
        "sha": sha,
    }
    response = _requests.put(url, headers=headers, json=payload, timeout=10)
    return response.status_code in (200, 201)


def get_next_codes(count):
    """품목코드 자동 생성 (AI-000001, AI-000002, ...)
    GitHub의 counter.json을 통해 모든 직원이 같은 카운터 공유.
    동시 접근 시 SHA 충돌이 나면 재시도.
    """
    max_retries = 5
    for attempt in range(max_retries):
        try:
            current, sha = _github_get_counter()
            new_current = current + count
            
            if _github_update_counter(new_current, sha):
                # 성공 → 코드 반환
                codes = []
                for i in range(count):
                    codes.append(f"AI-{(current + i):06d}")
                return codes
            else:
                # 충돌 (다른 직원이 동시에 업데이트) → 잠깐 대기 후 재시도
                _time.sleep(0.5 + attempt * 0.3)
        except Exception as e:
            if attempt == max_retries - 1:
                raise
            _time.sleep(0.5 + attempt * 0.3)
    
    # 모든 재시도 실패 (거의 불가능)
    raise RuntimeError("코드 발급 실패: 다른 직원이 동시에 다운로드 중일 수 있어요. 잠시 후 다시 시도해주세요.")


def peek_next_code():
    """다음에 부여될 코드 미리보기 (카운터 변경 안 함)
    GitHub에서 현재 값만 읽음.
    """
    try:
        current, _ = _github_get_counter()
        return current
    except Exception:
        # GitHub 통신 실패 시 임시값
        return 1


# ============================================================
# 매입처/약자 처리
# ============================================================

def normalize_supplier(supplier_raw, corp_models_list, canonical_names=None):
    """매입처 정규화: (주) 등록목록 매칭 + canonical 통일 표기 적용"""
    if not supplier_raw:
        return ""
    cleaned = str(supplier_raw).strip()
    no_corp = re.sub(r'\(주\)|㈜|주식회사', '', cleaned).strip()
    
    # 1. canonical_names에 매칭되는 정식 표기로 통일 (O.K 상사 → 오케이상사)
    if canonical_names:
        if cleaned in canonical_names:
            return canonical_names[cleaned]
        if no_corp in canonical_names:
            return canonical_names[no_corp]
    
    # 2. (주) 등록 목록 매칭
    for corp_model in corp_models_list:
        corp_no_prefix = re.sub(r'\(주\)|㈜', '', corp_model).strip()
        if corp_no_prefix == no_corp:
            return corp_model
    return no_corp

def find_brand_abbr(supplier_clean, brand_abbr_dict):
    """매입처 → 약자 매칭"""
    # 1. 정확히 일치
    if supplier_clean in brand_abbr_dict:
        return brand_abbr_dict[supplier_clean]
    
    # 2. (주) 빼고 일치
    no_corp = re.sub(r'\(주\)|㈜', '', supplier_clean).strip()
    if no_corp in brand_abbr_dict:
        return brand_abbr_dict[no_corp]
    
    # 3. 부분 매칭 (예: "피피엘 주식회사" → "피피엘")
    for brand_name, abbr in brand_abbr_dict.items():
        clean_brand = re.sub(r'\(주\)|㈜|주식회사', '', brand_name).strip()
        clean_supplier = re.sub(r'\(주\)|㈜|주식회사', '', supplier_clean).strip()
        if clean_brand and clean_supplier and (clean_brand in clean_supplier or clean_supplier in clean_brand):
            return abbr
    
    return ""  # 못 찾으면 빈값


# ============================================================
# 텍스트 정리 (자체 약자, 모델코드, 대괄호 등 제거)
# ============================================================

def clean_raw_text(text):
    """명세서 원본 텍스트에서 노이즈 제거"""
    if not text:
        return ""
    s = str(text)
    # 1. 자체 회사 약자/접두 제거: (BN), [P], (P) 등 (영문 1~5자 괄호)
    s = re.sub(r'\([A-Z]{1,5}\)', '', s)
    s = re.sub(r'\[[A-Z]{1,5}\]', '', s)
    # 2. 거래처 자체 코드 제거: [A-1], [C-1], [B-2] 등 (짧은 영문-짧은숫자 패턴)
    # 단, [3000K], [화이트/3000K] 같은 켈빈/색상 정보는 보호
    s = re.sub(r'\[[A-Za-z]{1,3}\-\d{1,3}\]', '', s)
    s = re.sub(r'\[\d{1,3}\-[A-Za-z]{1,3}\]', '', s)
    # 3. 모델 시리얼 코드 제거: (LUCOB-1), (EL-9913), (LUDM-1)
    s = re.sub(r'\([A-Z][A-Z0-9\-]+\)', '', s)
    # 4. 대괄호 내용 자동 분류 처리:
    #    - 색상/켈빈 정보 ([화이트/3000K]) → 대괄호만 빼고 내용 유지
    #    - 짧은 한글 ([신성] 같은 브랜드) → 대괄호만 빼고 내용 유지
    #    - 모델 코드 ([SSD-500S]) → 통째로 제거
    #    - 메모성 정보 ([단가인상(251205)]) → 통째로 제거
    def _bracket_handler(match):
        content = match.group(1).strip()
        # 색상/켈빈 정보 → 보존
        if re.search(
            r'\d{4}\s*[KkＫｋ]|화이트|블랙|백색|흑색|실버|골드|블루|핑크|레드|그레이|베이지|브라운|샌드|우드|전구색|전주백색|주백색|주광색|전주백|전구|주광|주백',
            content
        ):
            return content
        # 영문대문자 + 하이픈/숫자 조합의 모델 코드 → 제거
        if re.match(r'^[A-Z][A-Z0-9\-]*\d', content):
            return ''
        # 짧은 한글 (1~5자) → 보존 (브랜드/제조사 등)
        if re.match(r'^[가-힣]{1,5}$', content):
            return content
        # 그 외 (긴 텍스트, 메모성, 숫자 포함 등) → 제거
        return ''
    s = re.sub(r'\[([^\]]+)\]', _bracket_handler, s)
    # 4-2. 단독 날짜성 괄호 제거: (251205), (240315), (12345678) 등 숫자만 4~8자리
    s = re.sub(r'\(\d{4,8}\)', '', s)
    # 5. 끝부분 -숫자 패턴 제거: "에디슨꼬23-100", "주광-200" 등의 끝 -숫자
    # 단, 5W나 3000K 같은 건 보호. 단순 -숫자 패턴만 제거
    s = re.sub(r'\-\d{2,4}(?=\s|$|[/])', '', s)
    # 6. 장식 기호 제거
    s = re.sub(r'[★※●◆■□▶◀◎○]', '', s)
    # 7. 연속 공백 정리
    s = re.sub(r'\s+', ' ', s).strip()
    return s


# ============================================================
# 켈빈 ↔ 색상 변환
# ============================================================

def kelvin_to_color_name(kelvin):
    """켈빈값 → 색상명"""
    k = int(kelvin)
    if 2000 <= k <= 3000:
        return "전구색"
    elif k == 3500:
        return "전주백색"
    elif k == 4000:
        return "주백색"
    elif 5000 <= k <= 7000:
        return "주광색"
    return None

def normalize_lightcolor(text):
    """불빛색상 정규화 - 켈빈/색상명이 있는 경우만 처리, 없으면 빈 문자열"""
    if not text:
        return ""
    s = str(text).strip()
    
    color_names = ['전구색', '전주백색', '주백색', '주광색']
    
    # "전구"만 있으면 "전구색"
    if s == '전구':
        return '전구색'
    
    # 켈빈 추출
    kelvin_match = re.search(r'(\d{4})\s*[KkＫｋ]', s)
    # 색상명 추출
    color_match = None
    for c in color_names:
        if c in s:
            color_match = c
            break
    if not color_match and re.search(r'(?:^|[\s/(])전구(?:[\s/)]|$)', s):
        color_match = '전구색'
    
    if kelvin_match and color_match:
        return f"{color_match}({kelvin_match.group(1)}K)"
    elif kelvin_match:
        kelvin = int(kelvin_match.group(1))
        cname = kelvin_to_color_name(kelvin)
        if cname:
            return f"{cname}({kelvin}K)"
        return f"{kelvin}K"
    elif color_match:
        return color_match
    
    # 켈빈도 색상도 없으면 빈 문자열
    return ""


# ============================================================
# 분류 판별
# ============================================================

CATEGORY_KEYWORDS = {
    '펜던트': ['펜던트', '팬던트', 'P/D', 'PD'],
    '매입등': ['다운라이트', '매입등', '매입', 'D/L', 'DL'],
    '배선기구': ['조광기', '디머', '스위치형', '컨트롤러', '리모컨'],
    '램프': ['MR16', 'MR11', 'PAR16', 'PAR20', 'PAR30', 'PAR38',
             '벌브', '인찌구', '볼구', '촛대구',
             '에디슨', '필라멘트', '빔스틱', 'T벌브',
             'G45', 'G80', 'G95', 'G125', 'G9',
             'A19', 'A21', 'A60', 'A65', 'A70', 'A75', 'A80', 'A95', 'A110',
             'C35', 'B11', 'B35',
             'T3', 'T5', 'T8',
             'PL램프', 'PLC', 'PLD', 'PLT',
             '콘벌브', '램프'],
}

# 램프 사이즈 코드 → 한글 종류명
LAMP_SIZE_MAP = {
    'G45': '인찌구',
    'A19': '벌브', 'A21': '벌브', 'A60': '벌브', 'A65': '벌브', 'A70': '벌브',
    'A75': '벌브', 'A80': '벌브', 'A95': '벌브', 'A110': '벌브',
    'G80': '볼구', 'G95': '볼구', 'G125': '볼구',
    'C35': '촛대구',
}

# 소켓 패턴 (긴 것부터 매칭해야 GU5.3이 GU5로 안 잘림)
SOCKET_PATTERNS = ['GU5.3', 'GU10', 'GU24', 'GU4', 
                   'E11', 'E12', 'E14', 'E17', 'E26', 'E27', 'E40',
                   'B22', 'B15']

# 램프 형태 수식어 (분리 처리)
LAMP_MODIFIERS = ['에디슨', '필라멘트', '촛대', '미니', '빈티지']

# 모델 코드 추출용 정규식 (사전 안의 코드 패턴)
MODEL_CODE_PATTERN = re.compile(
    r'\b(G\d{2,3}|A\d{2,3}|MR\d{1,2}|PAR\d{2}|C\d{2}|B\d{1,2}|T\d{1,2}|PL[CDT]?\d*)\b'
)


def detect_special_category(text):
    """레일/마그네틱 특수 분류 판별 (와트/부속 여부 등으로 4가지 구분)"""
    if not text:
        return ''
    s = str(text)
    
    has_watt = bool(re.search(r'\d+(?:\.\d+)?\s*[Ww]\b', s))
    has_lightcolor = bool(re.search(
        r'전구색|전주백색|주백색|주광색|\d{4}\s*[KkＫｋ]|전구|주광|주백', s
    ))
    
    # 마그네틱 제품군 판별: '마그네틱' 단어가 있는 경우만
    # (M5, T20 같은 시리즈 코드만으로는 인식 안 함 - 모델코드와 충돌 방지)
    if '마그네틱' in s:
        # 마그네틱 A타입/B타입 → 마그네틱부속 (레일이지만)
        if re.search(r'마그네틱\s*[AaBb]\s*타입', s):
            return '마그네틱부속'
        # 부속/전원/입력선/연결선
        if re.search(r'부속|전원|입력선|연결선|선재|컨넥터|커넥터', s):
            return '마그네틱부속'
        # 와트와 불빛색 둘 다 있으면 마그네틱조명
        if has_watt and has_lightcolor:
            return '마그네틱조명'
        # 그 외는 부속
        return '마그네틱부속'
    
    # 레일 처리 (마그네틱 아닌 경우)
    if '레일' in s:
        # CCA레일 또는 레일 N M (1M, 2M, 3M 등) → 레일부속
        if re.search(r'CCA\s*레일|레일\s*\d+\s*[Mm]\b', s):
            return '레일부속'
        # 부속/전원/입력선/연결선
        if re.search(r'부속|전원|입력선|연결선|선재|컨넥터|커넥터', s):
            return '레일부속'
        # 와트와 불빛색 둘 다 있으면 레일등기구
        if has_watt and has_lightcolor:
            return '레일등기구'
        # 그 외는 부속
        return '레일부속'
    
    return ''


def detect_category(text):
    """품명에서 마지막에 나오는 분류 키워드 찾기"""
    if not text:
        return ''
    
    # 1순위: 레일/마그네틱 특수 분류
    special = detect_special_category(text)
    if special:
        return special
    
    s = str(text).upper()
    
    matches = []
    for category, keywords in CATEGORY_KEYWORDS.items():
        for kw in keywords:
            kw_upper = kw.upper()
            idx = s.rfind(kw_upper)
            if idx >= 0:
                matches.append((idx, category, kw))
    
    # OCR 변형 패턴 추가: P/숫자/D, P숫자D 등 → 펜던트
    pd_variant = re.search(r'P\s*/?\s*\d*\s*/?\s*D', s)
    if pd_variant:
        matches.append((pd_variant.start(), '펜던트', 'P/D-variant'))
    
    # OCR 변형 패턴 추가: D/숫자/L, D숫자L 등 → 매입등
    dl_variant = re.search(r'D\s*/?\s*\d*\s*/?\s*L', s)
    if dl_variant:
        matches.append((dl_variant.start(), '매입등', 'D/L-variant'))
    
    if matches:
        matches.sort(key=lambda x: x[0], reverse=True)
        return matches[0][1]
    return ''


def extract_lamp_info(text):
    """램프 텍스트에서 모델코드, 소켓, 수식어 추출"""
    info = {
        'model_code': '',     # G45, MR16 등
        'korean_name': '',    # 인찌구, 벌브 등
        'socket': '',         # GU10, E26 등
        'modifier': '',       # 에디슨 등
        'tricolor': False,    # 삼색변환
    }
    
    # 1. 모델 코드 추출 (긴 것부터 매칭)
    text_upper = text.upper()
    
    # 사전에 있는 코드 우선
    for code, kname in sorted(LAMP_SIZE_MAP.items(), key=lambda x: -len(x[0])):
        if re.search(r'\b' + re.escape(code) + r'\b', text_upper):
            info['model_code'] = code
            info['korean_name'] = kname
            break
    
    # 사전에 없으면 일반 패턴으로 추출
    if not info['model_code']:
        m = MODEL_CODE_PATTERN.search(text_upper)
        if m:
            info['model_code'] = m.group(1)
            # MR, PAR, T 등은 한글명 없음 → model_code 자체가 램프이름
    
    # 2. 소켓 추출
    for socket in sorted(SOCKET_PATTERNS, key=lambda x: -len(x)):
        if re.search(r'\b' + re.escape(socket) + r'\b', text_upper):
            info['socket'] = socket
            break
    
    # 3. 수식어 추출
    for mod in LAMP_MODIFIERS:
        if mod in text:
            info['modifier'] = mod
            break
    
    # 4. 삼색변환
    if '삼색변환' in text or '삼색' in text or '3색' in text:
        info['tricolor'] = True
    
    return info


# ============================================================
# 품목명 조립 (분류별 강제 포맷팅)
# ============================================================

def extract_features(raw_text, memo_text=""):
    """원본 텍스트에서 핵심 정보 추출 (코드 패턴)"""
    full_text = f"{raw_text} {memo_text}".strip()
    
    features = {
        'cob': False,
        'inch': '',
        'dimming': False,
        'waterproof': False,
        'focus_diffuse': '',  # 집중/확산
        'color': '',          # 화이트, 블랙, 백색 등
        'watt': '',           # 8W
        'lightcolor': '',     # 전구색(3000K)
        'led_required': False,  # W가 있으면 True
        'size_no': '',        # 1호, 2호 등 (펜던트 사이즈)
    }
    
    # COB
    if re.search(r'\bCOB\b', full_text, re.IGNORECASE):
        features['cob'] = True
    
    # 인치
    inch_match = re.search(r'(\d+)\s*인치', full_text)
    if inch_match:
        features['inch'] = f"{inch_match.group(1)}인치"
    
    # 디밍
    if '디밍' in full_text:
        features['dimming'] = True
    
    # 방습
    if '방습' in full_text:
        features['waterproof'] = True
    
    # 집중/확산
    if '집중' in full_text:
        features['focus_diffuse'] = '집중'
        features['cob'] = True  # 집중이면 COB 강제
    elif '확산' in full_text:
        features['focus_diffuse'] = '확산'
        features['cob'] = False  # 확산이면 COB 제거
    
    # 와트
    watt_match = re.search(r'(\d+(?:\.\d+)?)\s*[wW]\b', full_text)
    if watt_match:
        features['watt'] = f"{watt_match.group(1)}W"
        features['led_required'] = True  # W가 있으면 LED 필수
    
    # 색상 (간단한 패턴)
    color_keywords = ['화이트', '블랙', '백색', '흑색', '실버', '골드', '블루', '핑크', 
                       '레드', '그레이', '베이지', '브라운', '샌드화이트', '우드']
    for ck in color_keywords:
        if ck in full_text:
            features['color'] = ck
            break
    
    # N호 사이즈 (펜던트 등 사이즈 표기) - 손글씨 OCR 변형 포함
    # 예: 1호, 2호, 10호 / OCR 오류: 1로, 1오, 1토 등
    size_match = re.search(r'(\d+)\s*[호로오토]\b', full_text)
    if size_match:
        features['size_no'] = f"{size_match.group(1)}호"
    else:
        # P/D 약자 사이에 숫자가 끼어든 경우 (예: P/1D = P/D 1호)
        # 한글에 바로 붙어있어도 잡히도록 처리
        pd_with_num = re.search(r'(?<![A-Za-z0-9])P\s*/?\s*(\d+)\s*/?\s*D(?![A-Za-z0-9])', full_text, re.IGNORECASE)
        if pd_with_num:
            features['size_no'] = f"{pd_with_num.group(1)}호"
    
    # 펜던트로 추정되는 경우 (P/D가 있거나 펜던트 키워드) + 끝에 단독 숫자가 남아있으면 size_no로 추정
    is_pendant = bool(re.search(r'펜던트|팬던트|P\s*/?\s*\d*\s*/?\s*D', full_text, re.IGNORECASE))
    if is_pendant and not features.get('size_no'):
        # 끝부분의 단독 숫자 (1~3자리, 와트 표기는 W가 붙으니 제외됨)
        trail_match = re.search(r'(\d{1,3})\s*$', full_text.strip())
        if trail_match:
            features['size_no'] = f"{trail_match.group(1)}호"
    
    # 불빛색상
    features['lightcolor'] = normalize_lightcolor(full_text)
    
    return features


def remove_keywords_from_text(text, keywords_to_remove):
    """텍스트에서 키워드들 제거"""
    s = text
    for kw in keywords_to_remove:
        # 대소문자 구분 없이 제거
        s = re.sub(re.escape(kw), '', s, flags=re.IGNORECASE)
    s = re.sub(r'\s+', ' ', s).strip()
    return s


def extract_product_name(raw_text, features, category):
    """원본에서 상품명만 추출 (스펙 정보 다 빼고)"""
    s = clean_raw_text(raw_text)
    
    # 1. 켈빈괄호 제거: (5700K), (3000K)
    s = re.sub(r'\(\d{4}\s*[KkＫｋ]\)', '', s)
    # 2. 켈빈 단독 제거: 5700K, 3000K
    s = re.sub(r'\d{4}\s*[KkＫｋ]', '', s)
    # 3. 색상명 제거
    for cn in ['전구색', '전주백색', '주백색', '주광색', '전구', '주광', '주백', '전주백']:
        s = s.replace(cn, '')
    # 4. N호 사이즈 제거 (예: 1호, 2호) - 손글씨 OCR 변형 포함
    s = re.sub(r'\d+\s*[호로오토]\b', '', s)
    
    # 4-2. 펜던트 분류면 끝 단독 숫자(1~3자리)도 제거 (size_no로 처리됨)
    if category == '펜던트':
        s = re.sub(r'\s\d{1,3}\s*$', '', s)
    # 5. 분류 약자를 한글로 치환 (P/D → 펜던트, D/L → 다운라이트)
    # 한글에 바로 붙어있어도 잡히도록 lookahead/lookbehind 사용
    s = re.sub(r'(?<![A-Za-z0-9])P\s*/?\s*\d*\s*/?\s*D(?![A-Za-z0-9])', ' 펜던트 ', s, flags=re.IGNORECASE)
    s = re.sub(r'(?<![A-Za-z0-9])D\s*/?\s*\d*\s*/?\s*L(?![A-Za-z0-9])', ' 다운라이트 ', s, flags=re.IGNORECASE)
    
    # 6. 제거할 스펙 키워드들
    remove_kw = ['COB', '디밍', '방습', '집중', '확산', 'LED']  # LED도 제거 (코드가 다시 붙임)
    if features['inch']:
        remove_kw.append(features['inch'])
    if features['watt']:
        remove_kw.append(features['watt'])
    if features['color']:
        remove_kw.append(features['color'])
    
    s = remove_keywords_from_text(s, remove_kw)
    
    # 7. 빈 괄호/대괄호 제거
    s = re.sub(r'\(\s*\)', '', s)
    s = re.sub(r'\[\s*\]', '', s)
    
    # 8. 슬래시 잔재 정리 (앞뒤, 연속 슬래시)
    s = re.sub(r'/+', '/', s)
    s = re.sub(r'^[\s/]+', '', s)
    s = re.sub(r'[\s/]+$', '', s)
    s = re.sub(r'\s*/\s*', '/', s)
    
    # 9. 끝부분 잔재 제거
    s = re.sub(r'[\-_/\s]+$', '', s)
    
    # 10. 연속 공백 정리
    s = re.sub(r'\s+', ' ', s).strip()
    
    return s


def build_product_name(abbr, raw_text, memo_text, category, model_code_full=''):
    """분류별 품목명 조립"""
    features = extract_features(raw_text, memo_text)
    if model_code_full:
        features['model_code_full'] = model_code_full
    product = extract_product_name(raw_text, features, category)
    
    # 약자 prefix
    prefix = f"{abbr}) " if abbr else ") "
    
    # LED 필요 여부 (와트 있으면 무조건 LED)
    led_str = "LED" if features['led_required'] else ""
    
    if category == '매입등':
        # [약자]) LED [COB] [인치] [디밍] [방습] [상품명+형태]/[집중/확산]/[색상]/[W]/[불빛색]
        parts_front = []
        if led_str:
            parts_front.append(led_str)
        if features['cob']:
            parts_front.append('COB')
        if features['inch']:
            parts_front.append(features['inch'])
        if features['dimming']:
            parts_front.append('디밍')
        if features['waterproof']:
            parts_front.append('방습')
        parts_front.append(product)
        front = ' '.join(p for p in parts_front if p)
        
        # 슬래시 부분
        slash_parts = []
        if features['focus_diffuse']:
            slash_parts.append(features['focus_diffuse'])
        if features['color']:
            slash_parts.append(features['color'])
        if features['watt']:
            slash_parts.append(features['watt'])
        if features['lightcolor']:
            slash_parts.append(features['lightcolor'])
        
        if slash_parts:
            return f"{prefix}{front}/{'/'.join(slash_parts)}"
        return f"{prefix}{front}"
    
    elif category == '램프':
        # [약자]) LED [디밍] [램프이름]/[모델코드]/[W]/[소켓]/[불빛색]
        # - 한글명 있음: 램프이름 = "[수식어] [한글명]" / 모델코드 별도
        # - 한글명 없음: 램프이름 = 모델코드 (한 번만 표기)
        
        full_text = f"{raw_text} {memo_text}"
        lamp_info = extract_lamp_info(full_text)
        
        # 램프이름 구성
        if lamp_info['korean_name']:
            # 한글명 있음: [수식어] [한글명]
            name_parts = []
            if lamp_info['modifier']:
                name_parts.append(lamp_info['modifier'])
            name_parts.append(lamp_info['korean_name'])
            lamp_name = ' '.join(name_parts)
            model_code = lamp_info['model_code']
        else:
            # 한글명 없음: 모델코드만 (수식어가 있으면 앞에)
            if lamp_info['modifier']:
                lamp_name = f"{lamp_info['modifier']} {lamp_info['model_code']}"
            else:
                lamp_name = lamp_info['model_code']
            model_code = ''  # 중복 안 되게 비움
        
        # 앞부분 조립: LED + 디밍(있으면) + 램프이름
        parts_front = []
        if features['led_required'] or lamp_info['model_code']:
            parts_front.append('LED')
        if features['dimming']:
            parts_front.append('디밍')
        if lamp_name:
            parts_front.append(lamp_name)
        front = ' '.join(p for p in parts_front if p)
        
        # 슬래시 뒷부분 조립: [모델코드]/[W]/[소켓]/[불빛색]
        slash_parts = []
        if model_code:  # 한글명 있을 때만 모델코드 별도
            slash_parts.append(model_code)
        if features['watt']:
            slash_parts.append(features['watt'])
        if lamp_info['socket']:
            slash_parts.append(lamp_info['socket'])
        
        # 불빛색 + 삼색변환
        light_part = features['lightcolor']
        if lamp_info['tricolor']:
            light_part = f"{light_part}/삼색변환" if light_part else "삼색변환"
        if light_part:
            slash_parts.append(light_part)
        
        if slash_parts:
            return f"{prefix}{front}/{'/'.join(slash_parts)}"
        return f"{prefix}{front}"
    
    elif category == '펜던트':
        # [약자]) LED [상품명]/[색상]/[W]/[불빛색]
        # 일체형(W 있음) → LED·W 포함 / 비일체형(W 없음) → LED·W 제거
        parts_front = []
        if features['led_required']:  # W 있으면 LED
            parts_front.append('LED')
        parts_front.append(product)
        front = ' '.join(p for p in parts_front if p)
        
        slash_parts = []
        if features['color']:
            slash_parts.append(features['color'])
        if features['watt']:
            slash_parts.append(features['watt'])
        if features['lightcolor']:
            slash_parts.append(features['lightcolor'])
        
        if slash_parts:
            return f"{prefix}{front}/{'/'.join(slash_parts)}"
        return f"{prefix}{front}"
    
    elif category == '배선기구':
        # [약자]) [브랜드/제조사] [상품명] [W]/[모델코드]
        # 예: OK) 신성 조광기 스위치형 500W/SSD-500S
        # 와트가 있으면 상품명 뒤에 공백으로, 모델코드는 슬래시 뒤
        front = product
        if features['watt']:
            front = f"{front} {features['watt']}".strip()
        
        if features.get('model_code_full'):
            return f"{prefix}{front}/{features['model_code_full']}"
        return f"{prefix}{front}"
    
    elif category in ('레일등기구', '마그네틱조명'):
        # 펜던트와 동일한 양식: [약자]) LED [상품명]/[색상]/[W]/[불빛색]
        parts_front = []
        if features['led_required']:
            parts_front.append('LED')
        parts_front.append(product)
        front = ' '.join(p for p in parts_front if p)
        
        slash_parts = []
        if features['color']:
            slash_parts.append(features['color'])
        if features['watt']:
            slash_parts.append(features['watt'])
        if features['lightcolor']:
            slash_parts.append(features['lightcolor'])
        
        if slash_parts:
            return f"{prefix}{front}/{'/'.join(slash_parts)}"
        return f"{prefix}{front}"
    
    elif category in ('레일부속', '마그네틱부속'):
        # 부속품은 단순 형식: [약자]) [상품명]
        # 모델코드 있으면 슬래시로 추가
        if features.get('model_code_full'):
            return f"{prefix}{product}/{features['model_code_full']}"
        return f"{prefix}{product}".strip()
    
    else:
        # 분류 미정 → 단순 형식
        return f"{prefix}{product}".strip()


def build_specification(features):
    """규격 컬럼: 색상/W/불빛색/N호 (모델코드 절대 포함 안 함)"""
    parts = []
    if features['color']:
        parts.append(features['color'])
    if features['watt']:
        parts.append(features['watt'])
    if features['lightcolor']:
        parts.append(features['lightcolor'])
    if features.get('size_no'):
        parts.append(features['size_no'])
    return '/'.join(parts)


# ============================================================
# UI: 로고 + 제목
# ============================================================

logo_base64 = get_image_base64("logo.png")
if logo_base64:
    st.markdown(f'<div style="display:flex;justify-content:center;margin-top:-5px;"><img src="data:image/png;base64,{logo_base64}" width="90"></div>', unsafe_allow_html=True)
st.markdown('<div class="main-title">노랑조명 명세서 「얼마에요」 엑셀 자동 변환기</div>', unsafe_allow_html=True)

icon_base64 = get_image_base64("icon1.png") 
icon_html = f'<img src="data:image/png;base64,{icon_base64}" width="120">' if icon_base64 else ""

st.markdown(f"""
    <div class="custom-label-wrapper" style="position: relative; height: 0; z-index: 10;">
        <div style="position: absolute; top: 110px; left: 0; right: 0; text-align: center; color: #444;">
            <p style="font-size: 21px; font-weight: 700; margin: 0;">명세서 이미지 파일을 여기에 끌어다 놓으세요</p>
            <p style="font-size: 15px; font-weight: 400; color: #888; margin: 10px 0 15px 0;">(클릭 후 파일 선택 가능)</p>
            <p style="font-size: 11px; font-weight: 400; color: #bbb; margin: 0;">최대 200MB까지, 선명한 사진(JPG, PNG)을 권장합니다</p>
            <div style="margin-top: 40px;">{icon_html}</div>
        </div>
    </div>
""", unsafe_allow_html=True)


# ============================================================
# AI 설정
# ============================================================

genai.configure(api_key=st.secrets["GEMINI_API_KEY"])
model = genai.GenerativeModel(
    'gemini-2.5-flash',
    generation_config={"response_mime_type": "application/json"}
)


# ============================================================
# 38컬럼 양식
# ============================================================

YALMA_COLUMNS = [
    '품목코드', '품목명', '품목유형', '매입단가', '매출단가', '기초재고량', '기초재고단가',
    '단위', '규격', '바코드', '주류품목여부', '주류용도구분', '주종구분', '분류명', '브랜드명',
    '모델명', '과세구분', '품목등록일자', '매입기준수량', '적정재고수량',
    '1등급가', '1등급수량', '2등급가', '2등급수량', '3등급가', '3등급수량',
    '4등급가', '4등급수량', '5등급가', '5등급수량',
    '사용상태', '재고계산여부', '원산지구분표시', '조달청식별코드', '참고사항',
    '전용창고코드', '전용창고명', '다공정 여부'
]


# ============================================================
# 메인 로직
# ============================================================

uploaded_files = st.file_uploader("", type=['jpg', 'jpeg', 'png'], accept_multiple_files=True)

# 새 파일 업로드 시 세션 초기화
if uploaded_files:
    current_files_id = ','.join([f"{f.name}_{f.size}" for f in uploaded_files])
    if st.session_state.get('current_files_id') != current_files_id:
        st.session_state.current_files_id = current_files_id
        st.session_state.excluded_files = set()  # 제외 목록 초기화
        for k in ['df_analyzed', 'all_suppliers', 'all_excluded', 
                  'all_empty_excluded', 'all_zero_excluded', 'excel_data', 'final_df']:
            if k in st.session_state:
                del st.session_state[k]

# 활성 파일 (제외 안 된 파일들)
active_files = []
if uploaded_files:
    excluded = st.session_state.get('excluded_files', set())
    active_files = [f for f in uploaded_files if f"{f.name}_{f.size}" not in excluded]

if active_files:
    # 업로드된 모든 이미지 미리보기
    if len(active_files) == 1:
        f = active_files[0]
        image = Image.open(f)
        st.image(image, caption=f.name, use_container_width=True)
        # 1장일 때는 가운데에 작은 제외 버튼
        col_a, col_b, col_c = st.columns([2, 1, 2])
        with col_b:
            if st.button("🗑️ 이미지 제외", key=f"del_single_{f.name}_{f.size}", use_container_width=True):
                st.session_state.excluded_files.add(f"{f.name}_{f.size}")
                st.rerun()
    else:
        st.markdown(f"### 업로드된 파일 {len(active_files)}장")
        cols = st.columns(min(len(active_files), 4))
        for idx, f in enumerate(active_files):
            with cols[idx % 4]:
                img = Image.open(f)
                st.image(img, caption=f.name, use_container_width=True)
                if st.button("🗑️ 제외", key=f"del_{f.name}_{f.size}", use_container_width=True):
                    st.session_state.excluded_files.add(f"{f.name}_{f.size}")
                    st.rerun()
elif uploaded_files:
    # 모두 제외된 상태
    excluded_count = len(st.session_state.get('excluded_files', set()))
    st.warning(f"⚠️ 업로드한 {excluded_count}장 모두 제외되었습니다.")
    if st.button("📌 제외 취소 (전체 다시 보기)", use_container_width=True):
        st.session_state.excluded_files = set()
        st.rerun()

if active_files:
    
    st.markdown("""
        <div style="
            background-color: #FFFDF0;
            border: 1px solid #FFD400;
            color: #8B7500;
            padding: 12px 15px;
            border-radius: 6px;
            text-align: center;
            font-size: 14px;
            margin: 10px 0;
            line-height: 1.6;
        ">
            사진이 흐리거나 글씨가 잘 안 보이면, 페이지를 새로고침(F5)하고 더 선명한 사진으로 다시 올려주세요.<br>
            <strong>※ 손글씨로 작성된 이미지는 입력이 잘 안될 수도 있습니다.</strong>
        </div>
    """, unsafe_allow_html=True)
    
    # ====================================================
    # 1단계: 분석 시작 (다중 파일 처리)
    # ====================================================
    if st.button('🚀 데이터 분석 시작', use_container_width=True):
        try:
            brand_data = load_brand_data()
            brand_abbr = brand_data['brand_abbreviations']
            corp_models = brand_data['models_with_corp_prefix']
            
            prompt = """
거래명세서 이미지를 분석해서 정보를 단순 추출해줘. 가공/포맷팅은 절대 하지 마.

## ⚠️ 매우 중요: 매입처 식별
거래명세서에는 보통 2개의 회사 정보가 있어:
- **공급자** (供給者) = 물건을 파는 쪽 = **매입처** ⭐ (이걸 supplier로 추출)
- **공급받는자** = 물건을 받는 쪽 = 우리 회사 (노랑조명) ❌ (이건 무시)

명세서에서 "공급자", "공급하는자", "공급자용" 영역의 회사명/상호를 supplier로 추출해.
"공급받는자", "수신자", "받는자" 영역의 회사명은 절대 supplier로 쓰지 마.

만약 둘 다 있으면 "공급자" 쪽이 정답. "노랑조명"이 보이면 그건 우리 회사니까 supplier 아님.

## 추출할 정보
1. **supplier**: 위 규칙대로 공급자(파는 쪽) 회사 상호
2. **각 행 정보**:
   - raw_name: 품명·규격 컬럼 텍스트 (그대로, 임의 변경 금지)
   - raw_memo: 적요/비고 컬럼 텍스트 (있으면 그대로, 없으면 빈 문자열)
   - 수량 (숫자만)
   - 단가 (부가세 미포함, 쉼표/원/통화기호 빼고 숫자만)
   - 단위 (있으면, 없으면 빈 문자열)

## ⚠️ 손글씨 명세서 단가 표기 (매우 중요)
손글씨 명세서에서는 천 단위 0 3개를 긴 가로선(-)으로 줄여 쓰는 경우가 많아:
- "40-" 또는 "40 -" → **40,000원** (단가: 40000)
- "120-" → 120,000원 (단가: 120000)
- "8-" → 8,000원 (단가: 8000)
- "2,800" 또는 "2800" → 2,800원 (단가: 2800, 그대로)
- "44-" → 44,000원 (단가: 44000)

규칙: 숫자 뒤에 가로선(-, 一, ㅡ)이 붙어있으면 → 1000을 곱해서 변환
숫자 뒤에 가로선이 없거나, 콤마/원 표기만 있으면 → 그대로

이 변환은 단가, 금액, 합계 모든 숫자 필드에 적용해.
JSON 응답의 "단가" 필드에는 변환된 최종 숫자(예: 40000)를 넣어줘.

## 절대 규칙
- 텍스트 임의 수정/요약/번역 금지
- 자체 약자((BN) 등) 제거 금지 (그대로 유지) — 코드에서 처리함
- 모델코드 (LUCOB-1) 등 제거 금지 (그대로 유지) — 코드에서 처리함
- 대괄호 [화이트/3000K] 그대로 유지 — 코드에서 처리함
- 매출단가 계산 금지 — 코드에서 처리함
- 분류 판별 금지 — 코드에서 처리함

## 출력 형식 (JSON만 응답)
{
  "supplier": "피피엘 주식회사",
  "items": [
    {
      "raw_name": "(BN)클레어 3인치 COB 매입 8W (LUCOB-1) [화이트/3000K]",
      "raw_memo": "★전구",
      "수량": 200,
      "단가": 4400,
      "단위": "EA"
    }
  ]
}

응답은 JSON 형식만, 다른 설명 절대 금지.
"""
            
            # 모든 명세서 결과 누적
            all_rows = []
            all_suppliers = []
            all_excluded = []
            all_empty_excluded = []
            all_zero_excluded = []
            failed_files = []
            
            # 진행 상황 표시
            progress_text = st.empty()
            progress_bar = st.progress(0)
            
            for idx, uploaded_file in enumerate(active_files):
                progress_text.markdown(f"**[{idx+1}/{len(active_files)}]** `{uploaded_file.name}` 분석 중...")
                progress_bar.progress((idx) / len(active_files))
                
                try:
                    image = Image.open(uploaded_file)
                    response = model.generate_content([prompt, image])
                    result = json.loads(response.text)
                    supplier_raw = result.get('supplier', '')
                    items_raw = result.get('items', [])
                    
                    if not items_raw:
                        failed_files.append(f"{uploaded_file.name} (품목 없음)")
                        continue
                    
                    supplier_clean = normalize_supplier(supplier_raw, corp_models, brand_data.get('canonical_names', {}))
                    
                    # 안전장치: 노랑조명 매입처 차단
                    if '노랑조명' in supplier_clean or '노랑' in supplier_clean.replace('(주)', ''):
                        failed_files.append(f"{uploaded_file.name} (매입처 잘못 인식: {supplier_raw})")
                        continue
                    
                    abbr = find_brand_abbr(supplier_clean, brand_abbr)
                    all_suppliers.append((supplier_clean, abbr, uploaded_file.name))
                    
                    # 부대비용 자동 제외
                    items_filtered = []
                    for item in items_raw:
                        raw_name = item.get('raw_name', '')
                        raw_memo = item.get('raw_memo', '')
                        if is_shipping_item(raw_name) or is_shipping_item(raw_memo):
                            all_excluded.append(f"[{supplier_clean}] {raw_name}")
                        else:
                            items_filtered.append(item)
                    
                    # 행 만들기
                    for item in items_filtered:
                        raw_name = item.get('raw_name', '')
                        raw_memo = item.get('raw_memo', '')
                        
                        model_code_full = ''
                        mc_match = re.search(r'\[([A-Z][A-Z0-9\-]*\d[A-Z0-9\-]*)\]', raw_name)
                        if mc_match:
                            model_code_full = mc_match.group(1)
                        
                        clean_name = clean_raw_text(raw_name)
                        clean_memo = clean_raw_text(raw_memo)
                        category = detect_category(clean_name + ' ' + clean_memo)
                        features = extract_features(clean_name, clean_memo)
                        features['model_code_full'] = model_code_full
                        item_name = build_product_name(abbr, clean_name, clean_memo, category, model_code_full)
                        spec = build_specification(features)
                        
                        # 빈 품목명 제외
                        if is_empty_item(item_name, abbr):
                            all_empty_excluded.append(f"[{supplier_clean}] {raw_name or '(빈 텍스트)'}")
                            continue
                        
                        purchase = pd.to_numeric(item.get('단가', 0), errors='coerce')
                        if pd.isna(purchase):
                            purchase = 0
                        purchase = int(purchase)
                        
                        # 단가 0원 제외
                        if purchase <= 0:
                            all_zero_excluded.append(f"[{supplier_clean}] {raw_name or '(빈 텍스트)'}")
                            continue
                        
                        sales = calculate_sales_price(purchase)
                        
                        row = {col: '' for col in YALMA_COLUMNS}
                        row['품목명'] = item_name
                        row['품목유형'] = '상품'
                        row['매입단가'] = purchase
                        row['매출단가'] = sales
                        row['단위'] = item.get('단위', '') or 'EA'
                        row['규격'] = spec
                        row['주류품목여부'] = '해당없음'
                        row['분류명'] = category
                        row['브랜드명'] = supplier_clean
                        row['모델명'] = supplier_clean
                        row['과세구분'] = '과세'
                        all_rows.append(row)
                
                except json.JSONDecodeError:
                    failed_files.append(f"{uploaded_file.name} (AI 응답 파싱 실패)")
                except Exception as e:
                    failed_files.append(f"{uploaded_file.name} (오류: {str(e)[:50]})")
            
            progress_bar.progress(1.0)
            progress_text.markdown(f"**✅ 분석 완료** ({len(active_files)}장 처리)")
            
            if not all_rows:
                st.error("❌ 처리된 품목이 없습니다. 사진을 더 선명하게 찍어 다시 시도해주세요.")
                if failed_files:
                    st.write("실패한 파일:", failed_files)
                st.stop()
            
            df = pd.DataFrame(all_rows, columns=YALMA_COLUMNS)
            for col in ['매입단가', '매출단가']:
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).astype(int)
            
            # 정렬: 매입처(브랜드명) → 분류명 → 품목명 가나다순
            df = df.sort_values(by=['브랜드명', '분류명', '품목명'], na_position='last').reset_index(drop=True)
            
            # 세션 저장 (코드 부여 전 상태)
            st.session_state.df_analyzed = df
            st.session_state.all_suppliers = all_suppliers
            st.session_state.all_excluded = all_excluded
            st.session_state.all_empty_excluded = all_empty_excluded
            st.session_state.all_zero_excluded = all_zero_excluded
            st.session_state.failed_files = failed_files
            if 'excel_data' in st.session_state:
                del st.session_state['excel_data']
        
        except Exception as e:
            st.error(f"❌ 분석 중 오류가 발생했습니다: {e}")
            import traceback
            with st.expander("자세한 오류 정보"):
                st.code(traceback.format_exc())
    
    # ====================================================
    # 2단계: 결과 표 + 검토
    # ====================================================
    if 'df_analyzed' in st.session_state:
        df = st.session_state.df_analyzed
        all_suppliers = st.session_state.get('all_suppliers', [])
        all_excluded = st.session_state.get('all_excluded', [])
        all_empty_excluded = st.session_state.get('all_empty_excluded', [])
        all_zero_excluded = st.session_state.get('all_zero_excluded', [])
        failed_files = st.session_state.get('failed_files', [])
        
        # 매입처별 품목 수 집계
        supplier_summary = df.groupby('브랜드명').size().to_dict()
        suppliers_text_parts = []
        no_abbr_suppliers = []
        seen = set()
        for supplier_clean, abbr, fname in all_suppliers:
            if supplier_clean in seen:
                continue
            seen.add(supplier_clean)
            count = supplier_summary.get(supplier_clean, 0)
            abbr_text = f"<strong>{abbr}</strong>" if abbr else "<span style='color:#c00;'>약자없음</span>"
            suppliers_text_parts.append(f"<strong>{supplier_clean}</strong>({abbr_text}, {count}개)")
            if not abbr:
                no_abbr_suppliers.append(supplier_clean)
        
        # 데이터 분석 시작 버튼 아래 40px 여백
        st.markdown("<div style='height: 40px;'></div>", unsafe_allow_html=True)
        
        # 통합 회색 박스: 분석 완료 + 제외 내역
        info_lines = [
            f"✅ <strong>분석 완료!</strong> 명세서 {len(all_suppliers)}장 / 총 품목 {len(df)}개"
        ]
        if suppliers_text_parts:
            info_lines.append(f"&nbsp;&nbsp;&nbsp;매입처: " + ', '.join(suppliers_text_parts))
        if all_excluded:
            info_lines.append(f"📦 부대비용 항목 {len(all_excluded)}개 자동 제외: {', '.join(all_excluded)}")
        if all_empty_excluded:
            info_lines.append(f"🔍 손글씨 인식 불가로 자동 제외 {len(all_empty_excluded)}개: {', '.join(all_empty_excluded)}")
        if all_zero_excluded:
            info_lines.append(f"💰 단가 0원으로 자동 제외 {len(all_zero_excluded)}개: {', '.join(all_zero_excluded)}")
        if failed_files:
            info_lines.append(f"❌ 분석 실패 {len(failed_files)}개: {', '.join(failed_files)}")
        
        info_html = '<br>'.join(info_lines)
        st.markdown(f"""
            <div style="
                background-color: #F5F5F5;
                border: 1px solid #E0E0E0;
                color: #555;
                padding: 12px 16px;
                border-radius: 6px;
                font-size: 14px;
                line-height: 1.6;
                margin: 8px 0;
            ">
                {info_html}
            </div>
        """, unsafe_allow_html=True)
        
        if no_abbr_suppliers:
            st.warning(f"⚠️ 약자 미등록 매입처: {', '.join(no_abbr_suppliers)} - 표에서 품목명 앞을 직접 입력하세요.")
        
        st.markdown("<div style='height: 25px;'></div>", unsafe_allow_html=True)
        st.info("💡 표에서 직접 수정 가능합니다. 매입단가를 변경하면 매출단가가 자동 재계산됩니다.  \n품목코드는 다운로드 확정 시 자동 부여됩니다 (AI-000001 형식). 다른 직원이 먼저 받으면 번호가 달라질 수 있습니다.")
        
        # 코드 미리보기 박스 (표 오른쪽 위)
        next_id = peek_next_code()
        start_code = f"AI-{next_id:06d}"
        end_code = f"AI-{next_id + len(df) - 1:06d}"
        # 1개일 때는 단일 코드, 여러 개일 때는 범위
        if len(df) == 1:
            code_display = start_code
        else:
            code_display = f"{start_code} ~ {end_code}"
        
        # 상단 여백 (지금의 2배)
        st.markdown("<div style='height: 44px;'></div>", unsafe_allow_html=True)
        
        # 표와 같은 너비(1100px) 컨테이너 안에서 우측 정렬
        st.markdown(f"""
            <div class="code-preview-wrapper" style="
                width: 100%;
                display: flex;
                justify-content: flex-end;
            ">
                <div class="code-preview-box" style="
                    background-color: #FFFFFF;
                    border: 1px solid #DDDDDD;
                    padding: 8px 14px;
                    text-align: center;
                    font-size: 13px;
                    line-height: 1.2;
                    min-width: 230px;
                ">
                    <div style="color: #888; margin: 0; padding: 0;">부여 예정 코드</div>
                    <div style="font-weight: 700; font-size: 15px; color: #333; margin-top: 2px;">
                        {code_display}
                    </div>
                </div>
            </div>
        """, unsafe_allow_html=True)
        
        # 박스와 표 사이 약간의 간격
        st.markdown("<div style='height: 8px;'></div>", unsafe_allow_html=True)
        
        # 표 순서: 품목명 → 단위 → 규격 → 매입단가 → 매출단가 → 분류명 → 브랜드명 → 모델명
        display_cols = ['품목명', '단위', '규격', '매입단가', '매출단가', '분류명', '브랜드명', '모델명']
        edited_display = st.data_editor(
            df[display_cols],
            use_container_width=True,
            num_rows="dynamic",
            key="editor",
            column_config={
                "매입단가": st.column_config.NumberColumn(format="%d원"),
                "매출단가": st.column_config.NumberColumn(format="%d원"),
            }
        )
        
        for col in display_cols:
            if col in edited_display.columns:
                df[col] = edited_display[col]
        
        df['매출단가'] = df['매입단가'].apply(calculate_sales_price)
        st.session_state.df_analyzed = df  # 수정사항 반영
        
        st.markdown("---")
        
        # ====================================================
        # 3단계: 확정 - 코드 부여 + 엑셀 생성
        # ====================================================
        if st.button('✅ 검토 완료 - 코드 부여하고 엑셀 만들기', use_container_width=True, type='primary'):
            try:
                # 품목코드 부여 (이 시점에 카운터 +N)
                codes = get_next_codes(len(df))
                df['품목코드'] = codes
                
                # 양식 파일에 데이터 채움
                template_path = "ERP Insert.xlsx"
                wb = load_workbook(template_path)
                ws = wb['품목 엑셀 업로드 양식']
                
                max_row = ws.max_row
                if max_row > 1:
                    for row_idx in range(2, max_row + 1):
                        for col_idx in range(1, len(YALMA_COLUMNS) + 1):
                            ws.cell(row=row_idx, column=col_idx).value = None
                
                for r_idx, row_data in enumerate(df.itertuples(index=False), start=2):
                    for c_idx, value in enumerate(row_data, start=1):
                        ws.cell(row=r_idx, column=c_idx, value=value)
                
                output = io.BytesIO()
                wb.save(output)
                
                st.session_state.excel_data = output.getvalue()
                st.session_state.final_df = df.copy()
                st.rerun()
                
            except FileNotFoundError:
                st.error(f"❌ 양식 파일 'ERP Insert.xlsx'를 찾을 수 없습니다. app.py와 같은 폴더에 두세요.")
            except Exception as e:
                st.error(f"❌ 엑셀 생성 중 오류: {e}")
                import traceback
                with st.expander("자세한 오류 정보"):
                    st.code(traceback.format_exc())
        
        # ====================================================
        # 4단계: 다운로드 (확정 후에만)
        # ====================================================
        if 'excel_data' in st.session_state:
            final_df = st.session_state.final_df
            first_code = final_df['품목코드'].iloc[0] if len(final_df) > 0 else 'N/A'
            last_code = final_df['품목코드'].iloc[-1] if len(final_df) > 0 else 'N/A'
            
            if len(final_df) == 1:
                st.success(f"🎉 코드 부여 완료! **{first_code}** (1개)")
            else:
                st.success(f"🎉 코드 부여 완료! **{first_code} ~ {last_code}** ({len(final_df)}개)")
            
            # 파일명: 매입처가 1곳이면 매입처명, 여러 곳이면 "여러건"
            unique_suppliers_for_name = list(set(final_df['브랜드명'].tolist()))
            if len(unique_suppliers_for_name) == 1:
                file_suffix = unique_suppliers_for_name[0]
            else:
                file_suffix = f"여러건_{len(unique_suppliers_for_name)}곳"
            
            st.download_button(
                label="📥 얼마에요 등록용 엑셀 다운로드",
                data=st.session_state.excel_data,
                file_name=f"노랑조명_명세서_{file_suffix}_변환결과.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                type='primary'
            )

st.markdown("<div style='border-top:1px solid #eee;margin-top:50px;padding-top:20px;text-align:center;font-size:11px;color:#bbb;'>Developed & Managed by <b>Eugene UG</b></div>", unsafe_allow_html=True)
